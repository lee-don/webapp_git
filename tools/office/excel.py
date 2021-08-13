import openpyxl

import openpyxl



def save_excel(target_list, output_file_name):
    """
    将数据写入xlsx文件
    """
    if not output_file_name.endswith('.xlsx'):
        output_file_name += '.xlsx'

    # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
    wb = openpyxl.Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active
    title_data = ('a', 'b', 'c', 'd', 'e', 'f')
    target_list.insert(0, title_data)
    rows = len(target_list)
    lines = len(target_list[0])
    for i in range(rows):
        for j in range(lines):
            ws.cell(row=i + 1, column=j + 1).value = target_list[i][j]

    # 保存表格
    wb.save(filename=output_file_name)

def read_excel(input_file_name):
    """
    从xlsx文件中读取数据
    """
    workbook = openpyxl.load_workbook(input_file_name)

    print(workbook)
    # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
    for sn in workbook.sheetnames:
        table = workbook.get_sheet_by_name(sn)
        rows = table.max_row
        cols = table.max_column

        for row in range(rows):
            for col in range(cols):
                data = table.cell(row + 1, col + 1).value
                print(data, end=' ')


if __name__ == '__main__':
    openpyxl_data = [
        ('我', '们', '在', '这', '寻', '找'),
        ('我', '们', '在', '这', '失', '去'),
        ('p', 'y', 't', 'h', 'o', 'n')
    ]
    output_file_name = 'openpyxl_file.xlsx'
    # save_excel(openpyxl_data, output_file_name)
    test = read_excel(output_file_name)









